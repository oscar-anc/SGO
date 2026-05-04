from strings import S
from theme import QSSA
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
import re
"""
Página 5: Generación de Anexos

Shown after GENERAR is pressed on pagePolicies or pageFinance.
Lets the user choose which additional documents to generate alongside
the Carta de Despacho:

  1. Manual de Procedimiento en Caso de Siniestros
      — same 12 optional checkboxes as the old PageManual dialog,
        plus a "Seleccionar todos" master checkbox.

  2. Garantías Particulares
      — one collapsible section per registered policy branch.
        Each section contains a QTextEdit with a minimal toolbar
        (bullet list and numbered list only).
        Produces one Word page per policy with a page break between them.

Navigation:
  ATRAS  → back to whatever page navigated here (pageFinance or pagePolicies)
  GENERAR DOCUMENTOS → generate Carta + selected annexes

Data retention:
  annexData dict in app.py stores selections across navigations.
  Keys for garantías: policy widgetId → {'branch': str, 'numero': str, 'html': str}
  Keys for manual:    'manual_items' → list of bool (one per checkbox)
                      'manual_yes'    → bool (radio Sí selected)
                      'guarantees_yes' → bool (radio Sí selected)
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QScrollArea,
    QLabel, QPushButton, QRadioButton, QButtonGroup,
    QCheckBox, QTextEdit, QSizePolicy, QFrame,
    QDialog, QLineEdit, QComboBox, QTreeWidget, QTreeWidgetItem,
    QHeaderView, QAbstractItemView, QSplitter, QGridLayout,
)
from PySide6.QtCore import Signal, Qt, QMimeData
from PySide6.QtGui  import QTextListFormat, QTextCursor
from widgets import NoNewlineLineEdit, CardWidget, EndorsementPolicyCard, InsuredGroupCard, EndorsementTableCard, GuaranteesCard, GuaranteesTableCard, CustomMessageBox
from helpers import resource_path
from svgs import SVG_EXCEL_IMPORT, SVG_DOWNLOAD_TEMPLATE, get_svg_download_template


# ── Manual item definitions (mirrors the old pageManual.py) ───────────────────
# Manual items loaded from strings.py — edit text there
MANUAL_ITEMS = [S[f'manual_item_{i:02d}'] for i in range(2, 14)]
ITEM_1_NAME  = S['manual_item_01']
ITEM_1_PAGES = 2


class CleanTextEdit(QTextEdit):
    """
    QTextEdit subclass that strips all external formatting on paste.

    When the user pastes from Word, a browser, or any other rich-text source,
    only the plain text is inserted so fonts, colors, sizes and spacing from
    the external source do not bleed into the document.
    The user then applies bullets or numbering manually via the toolbar.
    """
    def insertFromMimeData(self, source: QMimeData):
        """Override paste to always insert plain text only."""
        if source.hasText():
            # Insert plain text at the current cursor position — discards all
            # HTML, RTF, and image content from the clipboard source.
            self.insertPlainText(source.text())


class GuaranteesEditor(QWidget):
    """
    One collapsible editor section per policy branch for Garantías Particulares.

    Contains:
      - A header label showing the policy branch name
      - A CleanTextEdit with a minimal toolbar (bullet list, numbered list)
      - Collapse/expand toggle matching the CardWidget accordion pattern

    html property returns the current QTextEdit content as HTML for persistence.
    setHtml() restores previously saved content.
    """
    def __init__(self, branch_name, parent=None):
        super().__init__(parent)
        self._branch = branch_name
        self._buildUI()

    def _buildUI(self):
        outerLayout = QVBoxLayout(self)
        outerLayout.setContentsMargins(*QSSA['margins_annex_toolbar_outer'])
        outerLayout.setSpacing(QSSA['spacing_annex_toolbar_outer'])

        # ── Header bar ───────────────────────────────────────────────────────
        headerWidget = QWidget()
        headerWidget.setObjectName("CardHeader")
        headerWidget.setMinimumHeight(QSSA['annex_toolbar_min_height'])
        headerLayout = QHBoxLayout(headerWidget)
        headerLayout.setContentsMargins(*QSSA['margins_annex_toolbar_inner'])

        titleLabel = QLabel(S['guarantees_editor_title'].format(branch=self._branch))
        titleLabel.setObjectName("CardTitle")
        headerLayout.addWidget(titleLabel)
        headerLayout.addStretch()

        self._toggleBtn = QPushButton()
        self._toggleBtn.setObjectName("CardToggle")
        self._toggleBtn.setProperty('role', 'accordion')
        self._toggleBtn.setCheckable(True)
        self._toggleBtn.setChecked(True)  # expanded by default
        self._toggleBtn.setFlat(True)
        self._toggleBtn.setFixedSize(*QSSA['annex_toolbar_btn_size'])
        color = QSSA.get('card_header_toggle')
        self._toggleBtn.setIcon(CardWidget._chevronPixmap(expanded=True, color=color))
        self._toggleBtn.toggled.connect(self._onToggle)
        headerLayout.addWidget(self._toggleBtn)

        outerLayout.addWidget(headerWidget)

        # ── Body: toolbar + text editor ───────────────────────────────────────
        self._body = QWidget()
        self._body.setObjectName("CardBody")
        bodyLayout = QVBoxLayout(self._body)
        bodyLayout.setContentsMargins(*QSSA['margins_annex_body'])
        bodyLayout.setSpacing(QSSA['spacing_annex_body'])

        # Toolbar — icon buttons loaded from SVG files in imgs/
        # SVGs use currentColor which is resolved at load time using toolbar_button_text
        # from QSSA. Size is controlled by toolbar_btn_width/height in theme.
        toolbar = QHBoxLayout()
        toolbar.setSpacing(QSSA['spacing_annex_toolbar_items'])

        def _makeToolbarBtn(svg_fn, tooltip, callback):
            from PySide6.QtGui import QIcon, QPixmap
            from PySide6.QtCore import QByteArray
            from theme import QSSA as _AT
            icon_color = _AT.get('toolbar_button_text')
            btn_w      = int(_AT.get('toolbar_btn_width'))
            btn_h      = int(_AT.get('toolbar_btn_height'))
            icon_sz    = int(_AT.get('toolbar_btn_icon_size'))
            btn = QPushButton()
            btn.setObjectName('ToolbarButton')
            btn.setTooltip(tooltip)
            btn.setFixedSize(btn_w, btn_h)
            px = QPixmap()
            px.loadFromData(QByteArray(svg_fn(icon_color).encode('utf-8')), 'SVG')
            if not px.isNull():
                btn.setIcon(QIcon(px.scaled(
                    icon_sz, icon_sz,
                    Qt.KeepAspectRatio, Qt.SmoothTransformation
                )))
            btn.clicked.connect(callback)
            return btn

        self._btnBold      = _makeToolbarBtn(get_svg_bold,         S['toolbar_bold_tip'],      self._toggleBold)
        self._btnItalic    = _makeToolbarBtn(get_svg_italic,       S['toolbar_italic_tip'],    self._toggleItalic)
        self._btnUnderline = _makeToolbarBtn(get_svg_underline,    S['toolbar_underline_tip'], self._toggleUnderline)
        self._btnBullet    = _makeToolbarBtn(get_svg_list_bullet,  S['toolbar_list_tip'],      self._insertBulletList)
        self._btnClear     = _makeToolbarBtn(get_svg_clear_format, S['toolbar_clear_tip'],     self._clearListFormat)

        for btn in (self._btnBold, self._btnItalic, self._btnUnderline,
                    self._btnBullet, self._btnClear):
            toolbar.addWidget(btn)
        toolbar.addStretch()
        bodyLayout.addLayout(toolbar)


        # Instruction label
        instrLabel = QLabel(
            S["editor_paste_hint"]
        )
        instrLabel.setObjectName("InstructionLabel")
        instrLabel.setWordWrap(True)
        bodyLayout.addWidget(instrLabel)

        # The main editor — CleanTextEdit strips paste formatting
        self._editor = CleanTextEdit()
        self._editor.setMinimumHeight(QSSA['annex_editor_min_height'])
        self._editor.setAcceptRichText(True)
        bodyLayout.addWidget(self._editor)

        outerLayout.addWidget(self._body)

    # ── Toolbar actions ───────────────────────────────────────────────────────

    def _insertBulletList(self):
        """Apply bullet list to current paragraph/selection."""
        cursor = self._editor.textCursor()
        fmt    = QTextListFormat()
        fmt.setStyle(QTextListFormat.Style.ListDisc)
        fmt.setIndent(1)
        cursor.createList(fmt)
        self._editor.setFocus()

    def _toggleBold(self):
        """Toggle bold on current selection or at cursor position."""
        from PySide6.QtGui import QTextCharFormat, QFont
        cursor = self._editor.textCursor()
        # fontWeight() returns 400 (Normal) or 700 (Bold)
        is_bold = cursor.charFormat().fontWeight() >= 600
        if cursor.hasSelection():
            fmt = QTextCharFormat()
            fmt.setFontWeight(QFont.Weight.Normal if is_bold else QFont.Weight.Bold)
            cursor.mergeCharFormat(fmt)
        else:
            self._editor.setFontWeight(
                QFont.Weight.Normal if is_bold else QFont.Weight.Bold
            )
        self._editor.setFocus()

    def _toggleItalic(self):
        """Toggle italic on current selection or at cursor position."""
        from PySide6.QtGui import QTextCharFormat
        cursor = self._editor.textCursor()
        is_italic = cursor.charFormat().fontItalic()
        if cursor.hasSelection():
            fmt = QTextCharFormat()
            fmt.setFontItalic(not is_italic)
            cursor.mergeCharFormat(fmt)
        else:
            self._editor.setFontItalic(not is_italic)
        self._editor.setFocus()

    def _toggleUnderline(self):
        """Toggle underline on current selection or at cursor position."""
        from PySide6.QtGui import QTextCharFormat
        cursor = self._editor.textCursor()
        is_underline = cursor.charFormat().fontUnderline()
        if cursor.hasSelection():
            fmt = QTextCharFormat()
            fmt.setFontUnderline(not is_underline)
            cursor.mergeCharFormat(fmt)
        else:
            self._editor.setFontUnderline(not is_underline)
        self._editor.setFocus()

    def _clearListFormat(self):
        """
        Remove list formatting from the current paragraph.
        Also resets the block indent to 0 so the paragraph returns to the
        left margin — without this, Qt leaves a residual indent after
        removing the list, which looks like a stray tab stop.
        """
        from PySide6.QtGui import QTextBlockFormat
        cursor = self._editor.textCursor()
        block  = cursor.block()
        lst    = block.textList()
        if lst:
            lst.remove(block)
        # Reset indent on the block so no phantom spacing remains
        blockFmt = QTextBlockFormat()
        blockFmt.setIndent(0)
        blockFmt.setTextIndent(0.0)
        cursor.setBlockFormat(blockFmt)
        self._editor.setFocus()

    # ── Collapse / expand ─────────────────────────────────────────────────────

    def _onToggle(self, expanded: bool):
        self._body.setVisible(expanded)
        color = QSSA.get('card_header_toggle')
        self._toggleBtn.setIcon(
            CardWidget._chevronPixmap(expanded=expanded, color=color)
        )

    # ── Public API ────────────────────────────────────────────────────────────

    def getHtml(self):
        """Return editor content as HTML for persistence in annexData."""
        return self._editor.toHtml()

    def setHtml(self, html):
        """Restore previously saved HTML content into the editor."""
        if html:
            self._editor.setHtml(html)

    def isEmpty(self):
        """Return True if the editor has no meaningful text content."""
        return not self._editor.toPlainText().strip()

    @property
    def branchName(self):
        return self._branch


class PageAnnex(QWidget):
    """
    Página 5: Generación de Anexos.

    Shown after the user chooses to generate annexes from the CustomMessageBox
    that appears when GENERAR is pressed on pageFinance or pagePolicies.

    Signals:
      signalBack(str)     — emitted with the origin page name ('finance' or 'policies')
      signalGenerate      — emitted when GENERAR DOCUMENTOS is pressed
    """

    signalBack          = Signal(str)   # carries origin: 'finance' or 'policies'
    signalGenerate      = Signal()
    signalClearCesiones = Signal()        # emitted when Limpiar clears cesiones data

    def __init__(self):
        super().__init__()

        # Tracks which page opened this one so ATRAS goes back correctly
        self._origin = 'finance'

        # References to dynamic widgets rebuilt on each setupPage() call
        self._manualCheckboxes  = []      # list of QCheckBox for manual items
        self._masterCheckbox    = None    # "Seleccionar todos" master checkbox
        self._guaranteeEditors  = {}      # { widgetId: GarantiasEditor }
        self._guaranteeCheckboxes   = {}      # { widgetId: QCheckBox }
        self._cesionesWidgets       = {}      # { policyKey: CesionesWidget }

        self.buildUI()

    # ── Static scaffold ───────────────────────────────────────────────────────

    def buildUI(self):
        """Build the permanent scaffold. Dynamic content inserted by setupPage()."""
        scrollArea = QScrollArea()
        scrollArea.setWidgetResizable(True)

        self.containerWidget = QWidget()
        self.mainLayout      = QVBoxLayout(self.containerWidget)
        self.mainLayout.addStretch()

        scrollArea.setWidget(self.containerWidget)

        # Navigation — fixed outside scrollArea so it is always visible
        self._btnBack    = QPushButton(S["nav_back"])
        self._btnClear   = QPushButton(S["nav_clear"])
        self._btnGenerar = QPushButton(S["nav_generate_docs"])
        self._btnBack.clicked.connect(self._onBack)
        self._btnClear.clicked.connect(self._onClearPage)
        self._btnGenerar.clicked.connect(self.signalGenerate.emit)

        navWidget = QWidget()
        navWidget.setObjectName("NavBar")
        navLayout = QHBoxLayout(navWidget)
        navLayout.setContentsMargins(*QSSA['margins_nav'])
        navLayout.addWidget(self._btnBack)
        navLayout.addStretch()
        navLayout.addWidget(self._btnClear)
        navLayout.addStretch()
        navLayout.addWidget(self._btnGenerar)

        pageLayout = QVBoxLayout(self)
        pageLayout.setContentsMargins(*QSSA['margins_zero'])
        pageLayout.setSpacing(QSSA['spacing_zero'])
        pageLayout.addWidget(scrollArea, 1)
        pageLayout.addWidget(navWidget)

    def _onClearPage(self):
        """Reset all selections on this page after user confirmation."""
        from PySide6.QtWidgets import QMessageBox
        parts = ['Manual', 'Garantías']
        if self._cesionesWidgets:
            parts.append('Cesiones de Derecho')
        body_msg = (
            '¿Está seguro que desea limpiar los datos de esta página?\n\n'
            f'Se perderán las selecciones de {", ".join(parts)}.'
        )
        reply = CustomMessageBox.question(
            self,
            S["dlg_clear_page_title"],
            body_msg,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.No:
            return
        # Reset manual radios
        if hasattr(self, '_manualNo'):
            self._manualNo.setChecked(True)
        if hasattr(self, '_manualCheckArea'):
            self._manualCheckArea.setVisible(False)
        for cb in self._manualCheckboxes:
            cb.setChecked(False)
        # Reset guarantees radios and editors
        if hasattr(self, '_guaranteesNo'):
            self._guaranteesNo.setChecked(True)
        if hasattr(self, '_guaranteesArea'):
            self._guaranteesArea.setVisible(False)
        for editor in self._guaranteeEditors.values():
            editor.setHtml('')
        for cb in self._guaranteeCheckboxes.values():
            cb.setChecked(False)

    def _onBack(self):
        self.signalBack.emit(self._origin)

    def _clearDynamic(self):
        """Remove all dynamic widgets above the stretch (last 1 item).
        Navigation is outside the scrollArea so it is no longer part of mainLayout."""
        # Block all signals before deletion to prevent ghost dialogs
        from PySide6.QtWidgets import QWidget as _QW
        for key, tableCard in self._cesionesWidgets.items():
            tableCard.blockSignals(True)
            for child in tableCard.findChildren(_QW):
                child.blockSignals(True)
        while self.mainLayout.count() > 1:
            item = self.mainLayout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
            elif item.layout():
                self._clearLayout(item.layout())
        self._manualCheckboxes      = []
        self._masterCheckbox        = None
        self._guaranteeEditors      = {}
        self._guaranteeCheckboxes   = {}
        self._cesionesWidgets       = {}   # { policyKey: EndorsementTableCard }

    def _clearLayout(self, layout):
        while layout.count():
            item = layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
            elif item.layout():
                self._clearLayout(item.layout())

    # ── Public setup ──────────────────────────────────────────────────────────

    def setupPage(self, policiesPage, annexData, endorsementData=None, origin='finance'):
        """
        Rebuild the dynamic section from current pagePolicies state and
        previously stored annexData.

        policiesPage: live PagePolicies instance to read policy branches from.
        annexData:    dict persisted by app.py across navigations.
        origin:       'finance' or 'policies' — determines ATRAS destination.
        """
        self._origin = origin
        self._clearDynamic()

        # ── Card 1: Manual de Procedimientos ─────────────────────────────────
        cardManual  = CardWidget(S["card_manual"],
                                 collapsible=False)
        manualInner = QVBoxLayout()

        # Yes/No radio buttons
        manualRadioRow = QHBoxLayout()
        manualRadioRow.setSpacing(QSSA['spacing_annex_radio_row'])
        self._manualGroup = QButtonGroup(self)
        self._manualYes   = QRadioButton("Sí")
        self._manualNo    = QRadioButton("No")
        self._manualGroup.addButton(self._manualYes)
        self._manualGroup.addButton(self._manualNo)
        manualRadioRow.addWidget(self._manualYes)
        manualRadioRow.addWidget(self._manualNo)
        manualRadioRow.addStretch()
        manualInner.addLayout(manualRadioRow)

        # Collapsible checkbox area
        self._manualCheckArea = QWidget()
        checkAreaLayout       = QVBoxLayout(self._manualCheckArea)
        checkAreaLayout.setSpacing(QSSA['spacing_annex_check_area'])
        checkAreaLayout.setContentsMargins(*QSSA['margins_annex_check_area'])

        # Master "Seleccionar todos" — standard checkbox, outside grid
        self._masterCheckbox = QCheckBox(S["manual_select_all"])
        self._masterCheckbox.setTristate(True)
        self._masterCheckbox.stateChanged.connect(self._onMasterToggled)
        checkAreaLayout.addWidget(self._masterCheckbox)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFrameShadow(QFrame.Sunken)
        checkAreaLayout.addWidget(sep)

        # Item 1 — always included, shown as disabled button full width
        item1_btn = QPushButton()
        item1_btn.setCheckable(True)
        item1_btn.setChecked(True)
        item1_btn.setEnabled(False)
        item1_btn.setProperty('role', 'manual-fixed')
        # Bold name + normal-weight (obligatorio) using rich-text-like approach
        # QPushButton doesn't support rich text so we set text only;
        # the QSS handles font-weight for the button text
        item1_btn.setText(f"{ITEM_1_NAME}  ({S.get('manual_item_obligatorio', 'obligatorio')})")
        checkAreaLayout.addWidget(item1_btn)

        # Items 2-13 — QPushButton checkable in 2-column grid
        savedItems = annexData.get('manual_items', [False] * len(MANUAL_ITEMS))
        gridWidget = QWidget()
        gridLayout = QGridLayout(gridWidget)
        gridLayout.setSpacing(QSSA['spacing_annex_check_area'])
        gridLayout.setContentsMargins(0, 0, 0, 0)
        for idx, name in enumerate(MANUAL_ITEMS):
            btn = QPushButton(name)
            btn.setCheckable(True)
            btn.setProperty('role', 'manual-item')
            if idx < len(savedItems):
                btn.setChecked(savedItems[idx])
            btn.toggled.connect(self._onItemToggled)
            row, col = divmod(idx, 2)
            gridLayout.addWidget(btn, row, col)
            self._manualCheckboxes.append(btn)
        checkAreaLayout.addWidget(gridWidget)

        manualInner.addWidget(self._manualCheckArea)
        self._manualCheckArea.setVisible(False)

        cardManual.setLayout(manualInner)

        # Restore radio state
        if annexData.get('manual_yes', False):
            self._manualYes.setChecked(True)
            self._manualCheckArea.setVisible(True)
            self._updateMasterState()
        else:
            self._manualNo.setChecked(True)

        # Wire radio signals AFTER restoring state to avoid spurious fires
        self._manualYes.toggled.connect(self._onManualAnswerChanged)
        self._manualNo.toggled.connect(self._onManualAnswerChanged)

        self.mainLayout.insertWidget(0, cardManual)

        # ── Card 2: Garantías Particulares ────────────────────────────
        cardGarantias  = CardWidget(S["card_guarantees"], collapsible=False)
        guaranteesInner = QVBoxLayout()

        # Yes/No radio buttons
        guaranteesRadioRow = QHBoxLayout()
        guaranteesRadioRow.setSpacing(QSSA['spacing_annex_radio_row'])
        self._guaranteesGroup = QButtonGroup(self)
        self._guaranteesYes   = QRadioButton("Sí")
        self._guaranteesNo    = QRadioButton("No")
        self._guaranteesGroup.addButton(self._guaranteesYes)
        self._guaranteesGroup.addButton(self._guaranteesNo)
        guaranteesRadioRow.addWidget(self._guaranteesYes)
        guaranteesRadioRow.addWidget(self._guaranteesNo)
        guaranteesRadioRow.addStretch()
        guaranteesInner.addLayout(guaranteesRadioRow)

        # Collapsible policy selection + editors area
        self._guaranteesArea = QWidget()
        gAreaLayout         = QVBoxLayout(self._guaranteesArea)
        gAreaLayout.setSpacing(QSSA['spacing_annex_g_area'])
        gAreaLayout.setContentsMargins(*QSSA['margins_annex_g_area'])

        # Build per-insured structure using GuaranteesTableCard
        allInsured      = [policiesPage.mainInsured] + policiesPage.additionalInsuredList
        multipleInsured = len(allInsured) > 1
        savedGarantias  = annexData.get('guarantees', {})

        for insured in allInsured:
            insuredName = insured.insuredName.text().strip() or S.get('placeholder_no_name', '(Sin nombre)')
            # Collect unique policies for this insured
            seen = set()
            insuredPolicies = []
            for policy in insured.policyList:
                wid = policy.widgetId
                if wid in seen:
                    continue
                seen.add(wid)
                branch = policy.policyBranch.currentText().strip()  # blank if no branch
                numero = getattr(policy, 'policyNum', None)
                numero = numero.text().strip() if numero else ''
                insuredPolicies.append((wid, branch, numero))

            # Wrap in InsuredGroupCard if multiple insured
            if multipleInsured:
                groupCard = InsuredGroupCard(insuredName)
                containerLayout = groupCard.bodyLayout()
                gAreaLayout.addWidget(groupCard)
            else:
                containerLayout = gAreaLayout

            multiPolicy = len(insuredPolicies) > 1
            # Show checkbox when: multiple insured OR this insured has multiple policies
            showCb = multipleInsured or multiPolicy
            for wid, branch, numero in insuredPolicies:
                card = GuaranteesTableCard(
                    branch_name=branch,
                    checked=(wid in savedGarantias),
                    show_checkbox=showCb
                )
                card._numero = numero
                if wid in savedGarantias:
                    card.setHtml(savedGarantias[wid].get('html', ''))
                containerLayout.addWidget(card)
                self._guaranteeEditors[wid]   = card
                self._guaranteeCheckboxes[wid] = card

        # Auto-expand InsuredGroupCard if any child card is active
        if multipleInsured:
            # Re-walk to find the groupCards and expand if needed
            # We stored cards in self._guaranteeEditors — find parent groupCards
            for i in range(gAreaLayout.count()):
                item = gAreaLayout.itemAt(i)
                if item and item.widget():
                    gc = item.widget()
                    if hasattr(gc, 'bodyLayout'):
                        bl = gc.bodyLayout()
                        any_active = any(
                            bl.itemAt(j).widget().isChecked()
                            for j in range(bl.count())
                            if bl.itemAt(j) and bl.itemAt(j).widget()
                            and hasattr(bl.itemAt(j).widget(), 'isChecked')
                        )
                        if any_active:
                            gc._toggleBtn.blockSignals(True)
                            gc._toggleBtn.setChecked(True)
                            gc._body.setVisible(True)
                            gc._toggleBtn.blockSignals(False)

        guaranteesInner.addWidget(self._guaranteesArea)
        self._guaranteesArea.setVisible(False)

        cardGarantias.setLayout(guaranteesInner)

        # Restore radio state
        if annexData.get('guarantees_yes', False):
            self._guaranteesYes.setChecked(True)
            self._guaranteesArea.setVisible(True)
        else:
            self._guaranteesNo.setChecked(True)

        # Wire radio signals AFTER restoring state
        self._guaranteesYes.toggled.connect(self._onGuaranteesAnswerChanged)
        self._guaranteesNo.toggled.connect(self._onGuaranteesAnswerChanged)

        self.mainLayout.insertWidget(1, cardGarantias)

        # ── Card 3: Cesiones de Derecho (only when Detallado selected) ────────
        endorsement_type = ''
        if hasattr(policiesPage, 'endorsementDetailed'):
            if policiesPage.endorsementDetailed.isChecked():
                endorsement_type = 'Detallado'

        if endorsement_type == 'Detallado':
            self._buildCesionesCard(
                policiesPage, annexData, endorsementData or {}
            )

    # ── Cesiones de Derecho card builder ─────────────────────────────────────

    def _buildCesionesCard(self, policiesPage, annexData, endorsementData):
        """Build Card 3 — Cesiones de Derecho using EndorsementTableCard."""
        import json
        try:
            with open(resource_path('db.json'), encoding='utf-8') as _f:
                _db = json.load(_f)
            endosatarios = _db.get('endosatarios', [])
        except Exception:
            endosatarios = []

        currency     = policiesPage.getData().get('currency', 'S/.')
        multiInsured = policiesPage.multipleInsured.isChecked()
        allInsured   = [policiesPage.mainInsured] + policiesPage.additionalInsuredList

        insuredGroups = []
        for insWidget in allInsured:
            insName = insWidget.insuredName.text() if multiInsured else ''
            pols = []
            for pol in insWidget.policyList:
                branch = pol.policyBranch.currentText()
                numero = pol.policyNum.text()
                key    = f"{insWidget.widgetId}::{pol.widgetId}"
                pols.append({'key': key, 'branch': branch, 'numero': numero,
                             'isTrec': branch.upper() == 'TREC'})
            insuredGroups.append({'insuredName': insName, 'policies': pols})

        totalPolicies   = sum(len(g['policies']) for g in insuredGroups)
        hasManyPolicies = totalPolicies > 1

        cardCesiones  = CardWidget(S['card_cesiones'], collapsible=False)
        cesionesOuter = QVBoxLayout()

        # Global action buttons at the top - right aligned
        globalBtnLayout = QHBoxLayout()
        globalBtnLayout.addStretch()

        self._btnGlobalImport = _makeSvgButton(
            SVG_EXCEL_IMPORT, size=32, icon_size=18,
            tooltip=S['cesiones_global_import'], role='endtable-import'
        )
        self._btnGlobalDownload = _makeSvgButton(
            SVG_DOWNLOAD_TEMPLATE, size=32, icon_size=18,
            tooltip=S['cesiones_global_download'], role='endtable-import'
        )
        self._btnGlobalImport.clicked.connect(self._onGlobalImport)
        self._btnGlobalDownload.clicked.connect(self._onGlobalDownload)

        # Uniform height
        btn_height = QSSA.get('endtable_topbar_height', 32)
        self._btnGlobalImport.setFixedHeight(btn_height)
        self._btnGlobalDownload.setFixedHeight(btn_height)

        globalBtnLayout.addWidget(self._btnGlobalImport)
        globalBtnLayout.addWidget(self._btnGlobalDownload)
        cesionesOuter.addLayout(globalBtnLayout)

        for grp in insuredGroups:
            insName = grp['insuredName']
            pols    = grp['policies']

            if multiInsured:
                _displayName = insName if insName else S['insured_no_name']
                groupCard = InsuredGroupCard(insuredName=_displayName)

            for pol in pols:
                key    = pol['key']
                branch = pol['branch']
                numero = pol['numero']
                isTrec = pol['isTrec']
                saved  = endorsementData.get(key, {})

                defaultCols = None  # EndorsementTableCard uses its own defaults
                hasData = bool(saved.get('rows') or saved.get('groups'))
                defaultChecked = saved.get('checked', hasData)

                tableCard = EndorsementTableCard(
                    label=branch,
                    isTrec=isTrec,
                    defaultColumns=defaultCols,
                    savedData=saved if saved else None,
                    showCheckbox=hasManyPolicies,
                    currency=currency,
                    branch=branch,
                    numero=numero,
                )
                tableCard._endosatarios = endosatarios
                tableCard._policyKey    = key

                if hasManyPolicies:
                    tableCard.setChecked(defaultChecked)

                if multiInsured:
                    groupCard.addPolicyCard(tableCard)
                else:
                    cesionesOuter.addWidget(tableCard)
                self._cesionesWidgets[key] = tableCard

            if multiInsured and hasManyPolicies:
                cesionesOuter.addWidget(groupCard)
            elif multiInsured and not hasManyPolicies:
                cesionesOuter.addWidget(groupCard)

        cardCesiones.setLayout(cesionesOuter)
        self.mainLayout.insertWidget(2, cardCesiones)

        # Auto-expand InsuredGroupCards that have active (checked + data) children
        if multiInsured:
            for i in range(cesionesOuter.count()):
                item = cesionesOuter.itemAt(i)
                if not item or not item.widget():
                    continue
                gc = item.widget()
                # InsuredGroupCard has _body and _toggleBtn
                if not (hasattr(gc, '_body') and hasattr(gc, '_toggleBtn')):
                    continue
                # Check if any child EndorsementTableCard is checked and has data
                bl = gc._body.layout() if gc._body.layout() else None
                if not bl:
                    continue
                any_active = False
                for j in range(bl.count()):
                    child_item = bl.itemAt(j)
                    if not child_item or not child_item.widget():
                        continue
                    w = child_item.widget()
                    cb_checked = (hasattr(w, '_cb') and w._cb is not None
                                  and w._cb.isChecked())
                    no_cb = (hasattr(w, '_cb') and w._cb is None)
                    has_data = hasattr(w, '_data') and bool(
                        w._data.get('rows') or w._data.get('groups'))
                    if (cb_checked or no_cb) and has_data:
                        any_active = True
                        break
                if any_active:
                    gc._toggleBtn.blockSignals(True)
                    gc._toggleBtn.setChecked(True)
                    gc._body.setVisible(True)
                    gc._toggleBtn.blockSignals(False)

    def _onGlobalImport(self):
        """Global import: reads Excel and distributes rows to matching policy cards."""
        import openpyxl
        from PySide6.QtWidgets import QFileDialog
        from widgets import CustomMessageBox

        path, _ = QFileDialog.getOpenFileName(
            self, S['cesiones_global_import_title'], '', 'Excel (*.xlsx *.xls)'
        )
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            rows_iter = list(ws.iter_rows(values_only=True))
            if not rows_iter:
                return
            excelCols = [str(c) if c is not None else '' for c in rows_iter[0]]
            excelRows = [[str(c) if c is not None else '' for c in row]
                         for row in rows_iter[1:] if any(c is not None for c in row)]
            
            # Group rows by policy using 'Ramo' and optionally 'Nro. Póliza' columns
            ramo_idx = None
            nro_idx = None
            for i, c in enumerate(excelCols):
                cl = c.lower()
                if 'ramo' in cl:
                    ramo_idx = i
                if any(k in cl for k in ['nro', 'numero', 'poliza']):
                    nro_idx = i
            
            # Build lookup: (ramo, numero) -> list of rows
            from collections import defaultdict
            grouped = defaultdict(list)
            for row in excelRows:
                ramo = row[ramo_idx].strip() if ramo_idx is not None and ramo_idx < len(row) else ''
                numero = row[nro_idx].strip() if nro_idx is not None and nro_idx < len(row) else ''
                grouped[(ramo, numero)].append(row)
            
            total_imported = 0
            for (ramo, numero), rows in grouped.items():
                # Find matching EndorsementTableCard
                for card_key, card in self._cesionesWidgets.items():
                    card_ramo = card._branch
                    card_numero = getattr(card, '_numero', '')
                    if card_ramo == ramo and (not numero or card_numero == numero):
                        # Import into this card
                        if card._isTrec:
                            card._importTrec(excelCols, rows)
                        else:
                            card._importStandard(excelCols, rows)
                        card._updateCountLabel()
                        total_imported += len(rows)
                        break
            
            CustomMessageBox.information(
                self, S['cesiones_global_import_title'],
                S['cesiones_global_import_ok'].format(n=total_imported)
            )
        except Exception as e:
            CustomMessageBox.critical(
                self, S['cesiones_global_import_title'],
                S['cesiones_global_import_err'].format(error=str(e))
            )

    def _onGlobalDownload(self):
        """Generate Excel template with all registered policies and insurers."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from PySide6.QtWidgets import QFileDialog
        from widgets import CustomMessageBox

        path, _ = QFileDialog.getSaveFileName(
            self, S['cesiones_global_download_title'],
            'plantilla_global_cesiones.xlsx', 'Excel (*.xlsx)'
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Plantilla Global'

            # Build headers from all registered policies
            headers = ['Asegurado', 'Ramo', 'Nro. Póliza']
            # Collect unique columns from all cards
            all_cols = set()
            for card_key, card in self._cesionesWidgets.items():
                if card._data.get('groups'):
                    for grp in card._data['groups']:
                        all_cols.update(grp.get('columns', []))
                else:
                    all_cols.update(card._data.get('columns', []))

            # Standard columns
            currency = getattr(self, '_currency', 'S/.')
            base_cols = ['Detalle', f'Monto Endosado {currency}', 'Endosatario']
            for c in base_cols:
                if c not in all_cols:
                    all_cols.add(c)
            extra_cols = sorted(all_cols - set(base_cols))
            headers.extend(base_cols[:2])  # Detalle, Monto
            headers.extend(extra_cols)
            if 'Endosatario' in all_cols:
                headers.append('Endosatario')

            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='217346')
                cell.alignment = Alignment(horizontal='center')

            # Add rows for each policy
            row_num = 2
            for card_key, card in self._cesionesWidgets.items():
                ws.cell(row=row_num, column=1, value='')
                ws.cell(row=row_num, column=2, value=card._branch)
                ws.cell(row=row_num, column=3, value=getattr(card, '_numero', ''))
                row_num += 1

            wb.save(path)
            CustomMessageBox.information(
                self, S['cesiones_global_download_title'],
                S['cesiones_global_download_ok'].format(path=path)
            )
        except Exception as e:
            CustomMessageBox.critical(
                self, S['cesiones_global_download_title'],
                S['cesiones_global_download_err'].format(error=str(e))
            )

    def _onManualAnswerChanged(self):
        """Show/hide the checkbox area based on Sí/No selection."""
        self._manualCheckArea.setVisible(self._manualYes.isChecked())

    def _onMasterToggled(self, state):
        """
        Master checkbox toggled by the user.
        stateChanged emits int in PySide6 — compare with enum .value explicitly.
        Checked (2) → check all items.
        Unchecked (0) → uncheck all items.
        PartiallyChecked (1) is set only programmatically; if the user clicks
        a partially-checked master it cycles to Checked — treat it as Checked.
        """
        checked_val       = Qt.CheckState.Checked.value       # 2
        partially_val     = Qt.CheckState.PartiallyChecked.value  # 1

        if state == partially_val:
            # Cycle to fully checked when user clicks an indeterminate master
            self._masterCheckbox.blockSignals(True)
            self._masterCheckbox.setCheckState(Qt.CheckState.Checked)
            self._masterCheckbox.blockSignals(False)
            state = checked_val

        target_checked = (state == checked_val)
        for cb in self._manualCheckboxes:
            cb.blockSignals(True)
            cb.setChecked(target_checked)
            cb.blockSignals(False)

    def _onItemToggled(self):
        """
        Update master checkbox state when an individual item is toggled.
        All checked   → master Checked.
        None checked  → master Unchecked.
        Some checked  → master PartiallyChecked (indeterminate — dash icon).
        """
        if self._masterCheckbox is None:
            return
        checked = [cb.isChecked() for cb in self._manualCheckboxes]
        self._masterCheckbox.blockSignals(True)
        if all(checked):
            self._masterCheckbox.setCheckState(Qt.CheckState.Checked)
        elif any(checked):
            self._masterCheckbox.setCheckState(Qt.CheckState.PartiallyChecked)
        else:
            self._masterCheckbox.setCheckState(Qt.CheckState.Unchecked)
        self._masterCheckbox.blockSignals(False)

    def _updateMasterState(self):
        """Sync master checkbox state to current individual items (used on restore)."""
        self._onItemToggled()

    # ── Garantías radio handler ───────────────────────────────────────────────

    def _onGuaranteesAnswerChanged(self):
        """Show/hide the policy editors area based on Sí/No selection."""
        self._guaranteesArea.setVisible(self._guaranteesYes.isChecked())

    # ── Data collection ───────────────────────────────────────────────────────

    def collectAnnexData(self):
        """
        Collect current selections and editor content into a dict for app.py.

        Returns:
        {
          'manual_yes':    bool,
          'manual_items': [bool, ...],   # one per MANUAL_ITEMS entry
          'guarantees_yes': bool,
          'guarantees':    { widgetId: {'branch': str, 'html': str}, ... }
        }
        """
        data = {}

        # Manual
        manual_si = (hasattr(self, '_manualYes') and self._manualYes.isChecked())
        data['manual_yes']    = manual_si
        data['manual_items'] = [cb.isChecked() for cb in self._manualCheckboxes]

        # Garantías
        garantias_si = (hasattr(self, '_guaranteesYes') and
                        self._guaranteesYes.isChecked())
        data['guarantees_yes'] = garantias_si

        garantias = {}
        # Only collect if user has selected Sí — prevents stale data
        # from persisting across navigation changes
        if garantias_si:
            for wid, card in self._guaranteeEditors.items():
                # For no-checkbox cards (single policy): only save if has content
                has_cb = getattr(card, '_showCheckbox', True)
                include = card.isChecked() if has_cb else not card.isEmpty()
                if include:
                    garantias[wid] = {
                        'branch': card.branchName,
                        'numero': getattr(card, '_numero', ''),
                        'html':   card.getHtml(),
                    }
        data['guarantees'] = garantias

        # Cesiones de Derecho
        cesiones = {}
        for key, tableCard in self._cesionesWidgets.items():
            wdata = tableCard.getData()
            # Only persist 'checked' when card has a real checkbox.
            # Without checkbox (single policy), omit 'checked' entirely
            # so rebuild uses hasData fallback instead of a stale value.
            has_cb = tableCard._cb is not None
            if has_cb:
                wdata['checked'] = tableCard._cb.isChecked()
            elif 'checked' in wdata:
                del wdata['checked']
            cesiones[key] = wdata
        data['cesiones'] = cesiones

        return data

    def getManualItems(self):
        """
        Return selected manual items with page numbering — same format as the
        old PageManual.getSelectedItems() so helpers.py needs no changes.
        Uses ITEM_1_PAGES defined at the top of this module — no pageManual import needed.
        """
        selected    = []
        currentPage = 1

        # Item 1 always included — uses module-level ITEM_1_PAGES constant
        selected.append({
            'number':    1,
            'name':      ITEM_1_NAME,
            'pages':     ITEM_1_PAGES,
            'startPage': currentPage,
        })
        currentPage += ITEM_1_PAGES

        # Optional items
        checkedItems = [
            (idx + 2, name, 2)
            for idx, (cb, name) in enumerate(
                zip(self._manualCheckboxes, MANUAL_ITEMS))
            if cb.isChecked()
        ]

        for i, (number, name, pages) in enumerate(checkedItems):
            isLast      = (i == len(checkedItems) - 1)
            actualPages = 3 if isLast else pages
            selected.append({
                'number':    number,
                'name':      name,
                'pages':     actualPages,
                'startPage': currentPage,
            })
            currentPage += actualPages

        return selected


# ──────────────────────────────────────────────────────────────────────
# CESIONES DE DERECHO WIDGETS
# ──────────────────────────────────────────────────────────────────────

def _makeSvgButton(svg_str, object_name='', size=32, icon_size=18, tooltip='', role=''):
    """
    Build a square icon-only QPushButton from an inline SVG string.
    Colors controlled via QSS role property (preferred) or objectName.
    """
    from PySide6.QtWidgets import QPushButton
    from PySide6.QtGui import QIcon, QPixmap
    from PySide6.QtCore import QByteArray, Qt

    btn = QPushButton()
    if role:
        btn.setProperty('role', role)
        btn.style().unpolish(btn)
        btn.style().polish(btn)
    elif object_name:
        btn.setObjectName(object_name)
    btn.setFixedSize(size, size)
    btn.setToolTip(tooltip)

    px = QPixmap()
    px.loadFromData(QByteArray(svg_str.encode('utf-8')), 'SVG')
    if not px.isNull():
        btn.setIcon(QIcon(px.scaled(
            icon_size, icon_size, Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation
        )))
    return btn


class CesionesWidget(QWidget):
    """
    Standard 3-column Cesiones entry widget for non-TREC policies.
    Layout:
        Row 1: QTextEdit (Detalle, full width, 4 lines high)
        Row 2: QLineEdit (Monto 30%) | QComboBox (Endosatario 70%)
        Row 3: Agregar | Actualizar | Eliminar  (equal width)
        Row 4: TreeWidget (3 cols, all stretch, col0 left, col1/2 center)
    """

    def __init__(self, key, endosatarios, currency, savedData=None, parent=None):
        super().__init__(parent)
        self._key      = key
        self._currency = currency
        self._countChangedCallback = None   # set by _buildCesionesCard
        self._buildUI(endosatarios, savedData or {})

    def _buildUI(self, endosatarios, saved):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(*QSSA['margins_annex_ces_card'])
        layout.setSpacing(QSSA['spacing_annex_ces_card'])

        # Row 1 — Detalle (full width, 4 lines ≈ 96px)
        self._detalle = QTextEdit()
        self._detalle.setPlaceholderText(S['cesiones_detalle_placeholder'])
        self._detalle.setFixedHeight(QSSA['annex_sig_preview_height'])
        self._detalle.setAcceptRichText(False)
        layout.addWidget(self._detalle)

        # Row 2 — Monto (30%) | Endosatario (70%)
        row2 = QHBoxLayout()
        self._monto = NoNewlineLineEdit()
        self._monto.setPlaceholderText(S['cesiones_monto_placeholder'])
        from validators import currencyInputValidator, applyCurrencyFormat, autoCompleter
        self._monto.setValidator(currencyInputValidator())
        self._monto.editingFinished.connect(lambda: applyCurrencyFormat(self._monto))
        self._endosatario = QComboBox()
        self._endosatario.setEditable(True)
        self._endosatario.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self._endosatario.addItem('')
        self._endosatario.addItems(endosatarios)
        self._endosatario.setCurrentText('')
        self._endosatario.setMinimumHeight(int(QSSA['combobox_min_height']))
        autoCompleter(self._endosatario, endosatarios)
        row2.addWidget(self._monto, 3)
        row2.addWidget(self._endosatario, 7)
        layout.addLayout(row2)

        # Row 3 — Buttons equal width
        row3 = QHBoxLayout()
        self._btnAdd    = QPushButton(S['btn_add'].upper())
        self._btnUpdate = QPushButton(S['btn_update'].upper())
        self._btnDelete = QPushButton(S['btn_delete'].upper())
        self._btnAdd.clicked.connect(self._onAdd)
        self._btnUpdate.clicked.connect(self._onUpdate)
        self._btnDelete.clicked.connect(self._onDelete)
        for b in (self._btnAdd, self._btnUpdate):
            b.setProperty('role', 'cesiones-neutral')
            b.style().unpolish(b); b.style().polish(b)
            row3.addWidget(b, 1)
        self._btnDelete.setProperty('role', 'cesiones-danger')
        self._btnDelete.style().unpolish(self._btnDelete); self._btnDelete.style().polish(self._btnDelete)
        row3.addWidget(self._btnDelete, 1)
        layout.addLayout(row3)

        # Row 4 — TreeWidget
        self._tree = QTreeWidget()
        self._tree.setColumnCount(3)
        self._tree.setHeaderLabels([
            S['cesiones_detalle_col'],
            f"{S['cesiones_monto_col']} {self._currency}",
            S['cesiones_endosatario_col'],
        ])
        self._tree.setIndentation(0)
        self._tree.setAlternatingRowColors(True)
        self._tree.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._tree.setMinimumHeight(QSSA['annex_sig_dialog_min_h'])
        # All columns stretch
        for c in range(3):
            self._tree.header().setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self._tree)

        # Restore rows — block ALL signals to prevent UI side effects during restore
        self._tree.blockSignals(True)
        for row in saved.get('rows', []):
            self._addTreeRow(row.get('detalle',''), row.get('monto',''), row.get('endosatario',''))
        self._tree.blockSignals(False)
        # Connect signals AFTER restore to prevent spurious triggers
        self._tree.itemSelectionChanged.connect(self._onSelectionChanged)
        self._tree.itemDoubleClicked.connect(self._onDoubleClick)

    def _addTreeRow(self, detalle, monto, endosatario):
        from PySide6.QtCore import Qt
        item = QTreeWidgetItem([detalle, monto, endosatario])
        item.setToolTip(0, detalle)
        item.setTextAlignment(1, Qt.AlignmentFlag.AlignCenter)
        item.setTextAlignment(2, Qt.AlignmentFlag.AlignCenter)
        self._tree.addTopLevelItem(item)

    def _onAdd(self):
        detalle     = self._detalle.toPlainText().strip()
        monto       = self._monto.text().strip()
        endosatario = self._endosatario.currentText().strip()
        if not detalle:
            return
        self._addTreeRow(detalle, monto, endosatario)
        self._detalle.clear()
        self._monto.clear()
        self._endosatario.setCurrentText('')
        self._notifyCount()

    def _onUpdate(self):
        from PySide6.QtCore import Qt
        item = self._tree.currentItem()
        if not item:
            return
        detalle     = self._detalle.toPlainText().strip()
        monto       = self._monto.text().strip()
        endosatario = self._endosatario.currentText().strip()
        if not detalle:
            return
        item.setText(0, detalle)
        item.setText(1, monto)
        item.setText(2, endosatario)
        item.setTextAlignment(1, Qt.AlignmentFlag.AlignCenter)
        item.setTextAlignment(2, Qt.AlignmentFlag.AlignCenter)
        item.setToolTip(0, detalle)

    def _onDelete(self):
        item = self._tree.currentItem()
        if not item:
            return
        self._tree.takeTopLevelItem(self._tree.indexOfTopLevelItem(item))
        self._notifyCount()

    def _onSelectionChanged(self):
        item = self._tree.currentItem()
        if item:
            self._detalle.setPlainText(item.text(0))
            self._monto.setText(item.text(1))
            self._endosatario.setCurrentText(item.text(2))

    def _onDoubleClick(self, item, col):
        if col != 0:
            return
        dlg = QDialog(self)
        dlg.setWindowTitle(S['cesiones_view_detail'])
        dlg.setMinimumWidth(QSSA['policy_editor_min_width'])
        dlgLayout = QVBoxLayout(dlg)
        te = QTextEdit()
        te.setPlainText(item.text(0))
        te.setReadOnly(True)
        te.setMinimumHeight(QSSA['policy_editor_min_height'])
        dlgLayout.addWidget(te)
        btnClose = QPushButton("Cerrar")
        btnClose.clicked.connect(dlg.accept)
        dlgLayout.addWidget(btnClose)
        dlg.exec()

    def _notifyCount(self):
        """Notify parent card of current row count for live update."""
        if callable(self._countChangedCallback):
            self._countChangedCallback(self._tree.topLevelItemCount(), False)

    def getRowCount(self):
        return self._tree.topLevelItemCount()

    def getData(self):
        rows = []
        for i in range(self._tree.topLevelItemCount()):
            item = self._tree.topLevelItem(i)
            rows.append({'detalle': item.text(0), 'monto': item.text(1), 'endosatario': item.text(2)})
        return {'rows': rows}


class TrecGroupWidget(CardWidget):
    """
    One collapsible group (bank/entity) with an 8-column TREC table.
    Layout inside:
        Row 1: QComboBox editable (entity name, full width) + trash icon btn
        Row 2: Equipo(70%) | Marca(30%)
        Row 3: Modelo(33%) | Serie(33%) | Placa(34%)
        Row 4: Año(33%) | Nro.Leasing(33%) | Suma Asegurada(34%) [numeric]
        Row 5: Agregar | Actualizar | Eliminar | [space] | ImportExcel | DescargarPlantilla
        Row 6: TreeWidget 8 cols all stretch, col0 left-aligned
    """

    def __init__(self, groupName='', endosatarios=None, currency='S/.', savedData=None, parent=None):
        super().__init__(title=groupName or 'Nueva Entidad', collapsible=True, parent=parent)
        self._currency    = currency
        self._endosatarios = endosatarios or []
        self._buildContent(groupName, savedData or {})

    def _buildContent(self, groupName, saved):
        from svgs import SVG_EXCEL_IMPORT, get_svg_trash
        from PySide6.QtCore import Qt
        inner = QVBoxLayout()
        inner.setSpacing(QSSA['spacing_annex_group_inner'])

        # Row 1 — Entity name combobox + trash remove button
        row1 = QHBoxLayout()
        self._nameCombo = QComboBox()
        self._nameCombo.setEditable(True)
        self._nameCombo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self._nameCombo.setPlaceholderText(S['trec_endosatario_ph'])
        self._nameCombo.addItems(self._endosatarios)
        self._nameCombo.setCurrentText(groupName)
        self._nameCombo.setMinimumHeight(int(QSSA['combobox_min_height']))
        from validators import autoCompleter
        autoCompleter(self._nameCombo, self._endosatarios)
        self._nameCombo.currentTextChanged.connect(
            lambda t: self.setTitle(t.strip() or 'Nueva Entidad')
        )
        self._btnRemove = _makeSvgButton(get_svg_trash(), 'BtnQuitar', tooltip='Quitar grupo')
        self._btnRemove.clicked.connect(self._onRemove)
        row1.addWidget(self._nameCombo, 1)
        row1.addWidget(self._btnRemove)
        inner.addLayout(row1)

        # Rows 2-4 — Input fields
        self._fEquipo  = NoNewlineLineEdit(); self._fEquipo.setPlaceholderText(S['trec_equipo_ph'])
        self._fMarca   = NoNewlineLineEdit(); self._fMarca.setPlaceholderText(S['trec_marca_ph'])
        self._fModelo  = NoNewlineLineEdit(); self._fModelo.setPlaceholderText(S['trec_modelo_ph'])
        self._fSerie   = NoNewlineLineEdit(); self._fSerie.setPlaceholderText(S['trec_serie_ph'])
        self._fPlaca   = NoNewlineLineEdit(); self._fPlaca.setPlaceholderText(S['trec_placa_ph'])
        self._fAnio    = NoNewlineLineEdit(); self._fAnio.setPlaceholderText(S['trec_anio_ph'])
        self._fLeasing = NoNewlineLineEdit(); self._fLeasing.setPlaceholderText(S['trec_leasing_ph'])
        self._fSuma    = NoNewlineLineEdit(); self._fSuma.setPlaceholderText(S['trec_suma_ph'])
        from validators import currencyInputValidator
        self._fSuma.setValidator(currencyInputValidator())

        row2 = QHBoxLayout()
        row2.addWidget(self._fEquipo,  7)
        row2.addWidget(self._fMarca,   3)
        inner.addLayout(row2)

        row3 = QHBoxLayout()
        row3.addWidget(self._fModelo, 333)
        row3.addWidget(self._fSerie,  333)
        row3.addWidget(self._fPlaca,  334)
        inner.addLayout(row3)

        row4 = QHBoxLayout()
        row4.addWidget(self._fAnio,    333)
        row4.addWidget(self._fLeasing, 333)
        row4.addWidget(self._fSuma,    334)
        inner.addLayout(row4)

        # Row 5 — Action buttons (equal) + spacer + icon buttons
        row5 = QHBoxLayout()
        row5.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        row5.setSpacing(QSSA['spacing_annex_row5'])
        self._btnAdd    = QPushButton(S['btn_add'].upper())
        self._btnUpdate = QPushButton(S['btn_update'].upper())
        self._btnDelete = QPushButton(S['btn_delete'].upper())
        self._btnAdd.clicked.connect(self._onAdd)
        self._btnUpdate.clicked.connect(self._onUpdate)
        self._btnDelete.clicked.connect(self._onDelete)
        for b in (self._btnAdd, self._btnUpdate):
            b.setProperty('role', 'cesiones-neutral')
            b.style().unpolish(b); b.style().polish(b)
            row5.addWidget(b, 1)
        self._btnDelete.setProperty('role', 'cesiones-danger')
        self._btnDelete.style().unpolish(self._btnDelete); self._btnDelete.style().polish(self._btnDelete)
        row5.addWidget(self._btnDelete, 1)
        row5.addStretch()
        self._btnImport = _makeSvgButton(SVG_EXCEL_IMPORT, size=26, icon_size=14, tooltip=S['btn_import_excel'], role='excel-icon')
        self._btnImport.clicked.connect(self._onImport)
        row5.addWidget(self._btnImport)
        inner.addLayout(row5)

        # Row 6 — TreeWidget 8 cols
        self._tree = QTreeWidget()
        self._tree.setColumnCount(8)
        self._tree.setHeaderLabels([
            S['trec_equipo_col'], S['trec_marca_col'], S['trec_modelo_col'],
            S['trec_serie_col'],  S['trec_placa_col'], S['trec_anio_col'],
            S['trec_leasing_col'], f"{S['trec_suma_col']} {self._currency}",
        ])
        self._tree.setIndentation(0)
        self._tree.setAlternatingRowColors(True)
        self._tree.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._tree.setMinimumHeight(QSSA['annex_tab_editor_min_h'])
        from PySide6.QtCore import Qt
        for c in range(8):
            self._tree.header().setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)
        # Header alignment via model: col0 left, rest center
        self._tree.model().setHeaderData(0, Qt.Orientation.Horizontal,
            int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter),
            Qt.ItemDataRole.TextAlignmentRole)
        for _c in range(1, 8):
            self._tree.model().setHeaderData(_c, Qt.Orientation.Horizontal,
                int(Qt.AlignmentFlag.AlignCenter),
                Qt.ItemDataRole.TextAlignmentRole)
        inner.addWidget(self._tree)

        self.setLayout(inner)

        # Restore rows — block signals to prevent selection handler firing
        self._tree.blockSignals(True)
        for row in saved.get('rows', []):
            self._addRow(row)
        self._tree.blockSignals(False)
        self._tree.itemSelectionChanged.connect(self._onSelectionChanged)

    _REMOVE_CALLBACK = None # set externally by TrecWidget

    def _onRemove(self):
        if callable(self._REMOVE_CALLBACK):
            self._REMOVE_CALLBACK(self)

    def _fieldValues(self):
        return [f.text() for f in (
            self._fEquipo, self._fMarca, self._fModelo, self._fSerie,
            self._fPlaca, self._fAnio, self._fLeasing, self._fSuma
        )]

    def _addRow(self, values):
        from PySide6.QtCore import Qt
        if isinstance(values, dict):
            keys = ('equipo','marca','modelo','serie','placa','anio','leasing','suma')
            values = [values.get(k,'') for k in keys]
        item = QTreeWidgetItem([str(v) if v is not None else '' for v in values])
        item.setTextAlignment(0, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        for c in range(1, 8):
            item.setTextAlignment(c, Qt.AlignmentFlag.AlignCenter)
        self._tree.addTopLevelItem(item)

    def _onAdd(self):
        vals = self._fieldValues()
        if not any(v.strip() for v in vals):
            return
        self._addRow(vals)
        for f in (self._fEquipo, self._fMarca, self._fModelo, self._fSerie,
                  self._fPlaca, self._fAnio, self._fLeasing, self._fSuma):
            f.clear()

    def _onUpdate(self):
        from PySide6.QtCore import Qt
        item = self._tree.currentItem()
        if not item:
            return
        for i, v in enumerate(self._fieldValues()):
            item.setText(i, v)
        item.setTextAlignment(0, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

    def _onDelete(self):
        item = self._tree.currentItem()
        if not item:
            return
        self._tree.takeTopLevelItem(self._tree.indexOfTopLevelItem(item))

    def _onSelectionChanged(self):
        item = self._tree.currentItem()
        if not item:
            return
        fields = (self._fEquipo, self._fMarca, self._fModelo, self._fSerie,
                  self._fPlaca, self._fAnio, self._fLeasing, self._fSuma)
        for i, f in enumerate(fields):
            f.setText(item.text(i))

    def _onImport(self):
        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getOpenFileName(self, S['cesiones_import_excel'], "", "Excel (*.xlsx *.xls)")
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            added = 0
            errors = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                vals = [str(c) if c is not None else '' for c in row[:8]]
                # Validate Suma Asegurada (col index 7) — reject if uses dot as thousands sep
                suma = vals[7] if len(vals) > 7 else ''
                if suma:
                    # Pattern: digit, dot, exactly 3 digits = thousands separator
                    if re.search(r'\d\.\d{3}', suma):
                        errors.append(S['cesiones_import_val_err'].format(row=row_idx, val=suma))
                        continue
                self._addRow(vals)
                added += 1
            msg = S['cesiones_import_ok'].format(n=added, g=1)
            if errors:
                msg += '\n\nAdvertencias:\n' + '\n'.join(errors)
                CustomMessageBox.warning(self, S['cesiones_import_excel'], msg)
            else:
                CustomMessageBox.information(self, S['cesiones_import_excel'], msg)
        except Exception as e:
            CustomMessageBox.critical(self, S['cesiones_import_excel'],
                                 S['cesiones_import_err'].format(error=str(e)))

    def _onDownload(self):
        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(self, S['cesiones_download_template'], "plantilla_trec.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = [
                'Equipo', 'Marca', 'Modelo', 'Serie', 'Placa',
                'Año', 'Nro. Leasing', f'Suma Asegurada {self._currency}'
            ]
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font    = Font(bold=True, color='FFFFFF')
                cell.fill    = PatternFill('solid', fgColor='217346')
                cell.alignment = Alignment(horizontal='center')
            # Add comment to Suma Asegurada column header
            suma_cell = ws.cell(row=1, column=8)
            comment = Comment(
                'Use punto (.) como separador decimal únicamente.\nNo use punto para separar miles (ej: 1000000.50, NO 1.000.000,50).',
                'Sistema'
            )
            suma_cell.comment = comment
            wb.save(path)
            CustomMessageBox.information(self, S['cesiones_download_template'],
                                    f'Plantilla guardada en:\n{path}')
        except Exception as e:
            CustomMessageBox.critical(self, S['cesiones_download_template'], str(e))

    def groupName(self):
        return self._nameCombo.currentText().strip()

    def getData(self):
        keys = ('equipo','marca','modelo','serie','placa','anio','leasing','suma')
        rows = []
        for i in range(self._tree.topLevelItemCount()):
            item = self._tree.topLevelItem(i)
            rows.append({k: item.text(j) for j, k in enumerate(keys)})
        return {'name': self.groupName(), 'rows': rows}


class TrecWidget(QWidget):
    """
    TREC mode: Import Excel + Download Template buttons at top,
    then Add Group button, then groups below.
    Initial state shows only the top buttons and Add Group.
    Groups are added by button or imported from Excel (grouped by Entidad column).
    """

    def __init__(self, key, endosatarios, currency, savedData=None, parent=None):
        super().__init__(parent)
        self._key                  = key
        self._endosatarios         = endosatarios
        self._currency             = currency
        self._groups               = []
        self._countChangedCallback = None   # set by _buildCesionesCard
        self._buildUI(savedData or {})

    def _buildUI(self, saved):
        self._outerLayout = QVBoxLayout(self)
        self._outerLayout.setContentsMargins(*QSSA['margins_annex_edit_outer'])
        self._outerLayout.setSpacing(QSSA['spacing_annex_edit_outer'])

        # Top action buttons: Import Excel | Download Template
        topRow = QHBoxLayout()
        topRow.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        topRow.setContentsMargins(*QSSA['margins_annex_edit_toprow'])
        topRow.setSpacing(QSSA['spacing_annex_edit_toprow'])
        self._btnImportAll   = QPushButton(S['btn_import_excel'])
        self._btnImportAll.setProperty('role', 'excel-text')
        self._btnImportAll.style().unpolish(self._btnImportAll)
        self._btnImportAll.style().polish(self._btnImportAll)
        self._btnImportAll.setFixedHeight(QSSA['policy_card_topbar_h'])
        self._btnDownloadAll = QPushButton(S['btn_download_template'])
        self._btnDownloadAll.setProperty('role', 'download')
        self._btnDownloadAll.style().unpolish(self._btnDownloadAll)
        self._btnDownloadAll.style().polish(self._btnDownloadAll)
        self._btnDownloadAll.setFixedHeight(QSSA['policy_card_topbar_h'])
        self._btnImportAll.clicked.connect(self._onImportAll)
        self._btnDownloadAll.clicked.connect(self._onDownloadTemplate)
        topRow.addStretch()
        topRow.addWidget(self._btnImportAll)
        topRow.addWidget(self._btnDownloadAll)
        topRow.addStretch()
        self._outerLayout.addLayout(topRow)

        # Groups container
        self._groupsContainer = QWidget()
        self._groupsLayout    = QVBoxLayout(self._groupsContainer)
        self._groupsLayout.setContentsMargins(*QSSA['margins_zero'])
        self._groupsLayout.setSpacing(QSSA['spacing_annex_groups_list'])
        self._outerLayout.addWidget(self._groupsContainer)

        # Add Group button
        self._btnAddGroup = QPushButton(S['btn_add_group'])
        self._btnAddGroup.clicked.connect(lambda: self._addGroup())
        self._outerLayout.addWidget(self._btnAddGroup)

        # Restore saved groups
        for gData in saved.get('groups', []):
            self._addGroup(gData.get('name',''), gData)

    def _addGroup(self, name='', savedData=None):
        grp = TrecGroupWidget(name, self._endosatarios, self._currency, savedData)
        grp._REMOVE_CALLBACK = self._removeGroup
        self._groups.append(grp)
        self._groupsLayout.addWidget(grp)
        self._notifyCount()

    def _removeGroup(self, grp):
        if grp in self._groups:
            self._groups.remove(grp)
        grp.deleteLater()
        self._notifyCount()

    def _onImportAll(self):
        """Import Excel grouped by Entidad column — creates TrecGroupWidgets automatically."""
        from PySide6.QtWidgets import QFileDialog
        import re
        path, _ = QFileDialog.getOpenFileName(self, S['cesiones_import_excel'], "", "Excel (*.xlsx *.xls)")
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            groups_data = {}
            group_order = []
            val_errors  = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(c is None for c in row):
                    continue
                vals = [str(c) if c is not None else '' for c in row[:9]]
                entidad = vals[0].strip()
                data_vals = vals[1:9] # cols 1-8: equipo..suma

                suma = data_vals[6] if len(data_vals) > 6 else ''
                if suma and re.search(r'\d\.\d{3}', suma):
                    val_errors.append(S['cesiones_import_val_err'].format(row=row_idx, val=suma))
                    continue

                if entidad not in groups_data:
                    groups_data[entidad] = []
                    group_order.append(entidad)
                groups_data[entidad].append(data_vals)

            total_rows = 0
            for entidad in group_order:
                rows_dicts = []
                keys = ('equipo','marca','modelo','serie','placa','anio','leasing','suma')
                for r in groups_data[entidad]:
                    rows_dicts.append({k: (r[i] if i < len(r) else '') for i, k in enumerate(keys)})
                self._addGroup(entidad, {'name': entidad, 'rows': rows_dicts})
                total_rows += len(rows_dicts)

            msg = S['cesiones_import_ok'].format(n=total_rows, g=len(group_order))
            if val_errors:
                msg += '\n\nAdvertencias:\n' + '\n'.join(val_errors)
                CustomMessageBox.warning(self, S['cesiones_import_excel'], msg)
            else:
                CustomMessageBox.information(self, S['cesiones_import_excel'], msg)
        except Exception as e:
            CustomMessageBox.critical(self, S['cesiones_import_excel'],
                                 S['cesiones_import_err'].format(error=str(e)))

    def _onDownloadTemplate(self):
        """Download Excel template with Entidad + 8 TREC columns."""
        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(self, S['cesiones_download_template'], "plantilla_trec.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = [
                'Entidad', 'Equipo', 'Marca', 'Modelo', 'Serie', 'Placa',
                'Año', 'Nro. Leasing', f'Suma Asegurada {self._currency}'
            ]
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font      = Font(bold=True, color='FFFFFF')
                cell.fill      = PatternFill('solid', fgColor='217346')
                cell.alignment = Alignment(horizontal='center')
            # Comment on Suma Asegurada
            suma_cell = ws.cell(row=1, column=9)
            suma_cell.comment = Comment(
                'Use punto (.) como separador decimal únicamente.\nNo use punto para separar miles (ej: 1000000.50, NO 1.000.000,50).',
                'Sistema'
            )
            wb.save(path)
            CustomMessageBox.information(self, S['cesiones_download_template'], f'Plantilla guardada en:\n{path}')
        except Exception as e:
            CustomMessageBox.critical(self, S['cesiones_download_template'], str(e))

    def _notifyCount(self):
        """Notify parent card of current group count for live update."""
        if callable(getattr(self, '_countChangedCallback', None)):
            self._countChangedCallback(len(self._groups), True)

    def getGroupCount(self):
        return len(self._groups)

    def getData(self):
        return {'groups': [g.getData() for g in self._groups]}
